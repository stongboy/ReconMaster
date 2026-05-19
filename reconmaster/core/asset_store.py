from __future__ import annotations

import csv
import io
import ipaddress
import json
import queue
import re
import socket
import sqlite3
import threading
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse


DOMAIN_RE = re.compile(r"(?i)\b(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}\b")
URL_RE = re.compile(r"(?i)https?://[^\s<>'\"]+")
IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
DEFAULT_COMPANY = "默认公司"


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


class PublicSuffix:
    def __init__(self, path: Path) -> None:
        self.exact: set[str] = set()
        self.wildcard: set[str] = set()
        self.exception: set[str] = set()
        self._load(path)

    def _load(self, path: Path) -> None:
        if not path.exists():
            return
        content = path.read_text(encoding="utf-8", errors="ignore")
        if content.lstrip().startswith("["):
            try:
                lines = json.loads(content)
            except json.JSONDecodeError:
                lines = []
        else:
            lines = content.splitlines()

        for raw in lines:
            line = str(raw).strip().lower()
            if not line or line.startswith("//"):
                continue
            if line.startswith("!"):
                self.exception.add(line[1:])
            elif line.startswith("*."):
                self.wildcard.add(line[2:])
            else:
                self.exact.add(line)

    def registrable_domain(self, host: str) -> str | None:
        labels = normalize_host(host).split(".")
        if len(labels) < 2 or any(not is_valid_label(label) for label in labels):
            return None

        best_suffix_len = 1
        for i in range(len(labels)):
            suffix = ".".join(labels[i:])
            if suffix in self.exception:
                best_suffix_len = max(1, len(labels) - i - 1)
                break
            if suffix in self.exact:
                best_suffix_len = max(best_suffix_len, len(labels) - i)
            if i + 1 < len(labels):
                wildcard_tail = ".".join(labels[i + 1 :])
                if wildcard_tail in self.wildcard:
                    best_suffix_len = max(best_suffix_len, len(labels) - i)

        if len(labels) <= best_suffix_len:
            return ".".join(labels)
        return ".".join(labels[-(best_suffix_len + 1) :])


class AssetStore:
    def __init__(self, db_path: Path, public_suffix_path: Path) -> None:
        self.db_path = db_path
        self.psl = PublicSuffix(public_suffix_path)
        self._lock = threading.RLock()
        self._resolve_queue: queue.Queue[int] = queue.Queue()
        self._init_db()
        self._resolver = threading.Thread(target=self._resolver_loop, daemon=True)
        self._resolver.start()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with self._lock, self._connect() as conn:
            conn.execute("PRAGMA journal_mode=WAL")
            legacy_name = self._rename_legacy_assets_if_needed(conn)
            self._create_schema(conn)
            self._ensure_company(conn, DEFAULT_COMPANY)
            if legacy_name:
                self._migrate_legacy_assets(conn, legacy_name)

    def _table_columns(self, conn: sqlite3.Connection, table: str) -> set[str]:
        return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}

    def _rename_legacy_assets_if_needed(self, conn: sqlite3.Connection) -> str | None:
        columns = self._table_columns(conn, "assets")
        if not columns:
            return None
        required = {"company_id", "company_name", "domain_key", "resolved_ip"}
        if required.issubset(columns):
            return None

        legacy_name = f"assets_legacy_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        for index in ("idx_assets_company", "idx_assets_root", "idx_assets_status", "idx_assets_source"):
            conn.execute(f"DROP INDEX IF EXISTS {index}")
        conn.execute(f"ALTER TABLE assets RENAME TO {legacy_name}")
        return legacy_name

    def _create_schema(self, conn: sqlite3.Connection) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                company_name TEXT NOT NULL,
                root_domain TEXT NOT NULL,
                subdomain TEXT,
                url TEXT,
                domain_key TEXT NOT NULL UNIQUE,
                resolved_ip TEXT,
                status TEXT NOT NULL DEFAULT 'Pending',
                source TEXT NOT NULL DEFAULT 'manual',
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                last_resolved_at TEXT,
                FOREIGN KEY(company_id) REFERENCES companies(id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fofa_query_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                query_text TEXT NOT NULL,
                processed_count INTEGER NOT NULL DEFAULT 0,
                include_domain INTEGER NOT NULL DEFAULT 1,
                include_cert INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS import_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_type TEXT NOT NULL,
                raw_content TEXT,
                total_count INTEGER DEFAULT 0,
                success_count INTEGER DEFAULT 0,
                failed_count INTEGER DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_company_id ON assets(company_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_company_name ON assets(company_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_root_domain ON assets(root_domain)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_status ON assets(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_source ON assets(source)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_import_records_created ON import_records(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_fofa_history_created ON fofa_query_history(created_at)")

    def _migrate_legacy_assets(self, conn: sqlite3.Connection, legacy_name: str) -> None:
        columns = self._table_columns(conn, legacy_name)
        rows = conn.execute(f"SELECT * FROM {legacy_name}").fetchall()
        for row in rows:
            data = dict(row)
            company = data.get("company") or data.get("company_name") or DEFAULT_COMPANY
            root = normalize_host(str(data.get("root_domain") or ""))
            subdomain = normalize_host(str(data.get("subdomain") or ""))
            url = normalize_url(str(data.get("url") or "")) if data.get("url") else ""
            if subdomain and not root:
                root = self.psl.registrable_domain(subdomain) or ""
            if root:
                root = self.psl.registrable_domain(root) or root
            if not root or not is_valid_domain(root):
                continue
            if subdomain == root:
                subdomain = ""
            if subdomain and not is_valid_domain(subdomain):
                subdomain = ""
            domain_key = build_domain_key(root, subdomain)
            company_id = self._ensure_company(conn, str(company))
            existing = conn.execute("SELECT id, source, url FROM assets WHERE domain_key = ?", (domain_key,)).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE assets
                    SET url = COALESCE(NULLIF(?, ''), url),
                        source = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        url,
                        merge_source(existing["source"], str(data.get("source") or "migration")),
                        now_iso(),
                        existing["id"],
                    ),
                )
                continue
            conn.execute(
                """
                INSERT INTO assets (
                    company_id, company_name, root_domain, subdomain, url, domain_key,
                    resolved_ip, status, source, notes, created_at, updated_at, last_resolved_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    str(company) or DEFAULT_COMPANY,
                    root,
                    subdomain,
                    url,
                    domain_key,
                    data.get("ip_address") if "ip_address" in columns else data.get("resolved_ip"),
                    data.get("status") or "Pending",
                    data.get("source") or "migration",
                    data.get("notes") or "",
                    data.get("created_at") or now_iso(),
                    data.get("updated_at") or now_iso(),
                    data.get("last_resolved_at"),
                ),
            )

    def _ensure_company(self, conn: sqlite3.Connection, name: str) -> int:
        company = clean_company(name)
        row = conn.execute("SELECT id FROM companies WHERE name = ?", (company,)).fetchone()
        if row:
            return int(row["id"])
        now = now_iso()
        conn.execute(
            "INSERT INTO companies (name, created_at, updated_at) VALUES (?, ?, ?)",
            (company, now, now),
        )
        return int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])

    def _company_for_root(self, conn: sqlite3.Connection, root_domain: str, fallback: str = "") -> str:
        row = conn.execute(
            """
            SELECT company_name FROM assets
            WHERE root_domain = ?
              AND (subdomain IS NULL OR subdomain = '' OR domain_key = ?)
            ORDER BY id ASC
            LIMIT 1
            """,
            (root_domain, root_domain),
        ).fetchone()
        if row and row["company_name"]:
            return str(row["company_name"])
        return clean_company(fallback)

    def preview(self, payload: dict[str, Any]) -> dict[str, Any]:
        mode = str(payload.get("mode", "single"))
        source = str(payload.get("source") or source_for_mode(mode))
        company = str(payload.get("company") or payload.get("company_name") or "")
        rows: list[dict[str, Any]] = []
        rejected: list[dict[str, str]] = []

        if mode == "single":
            company = clean_company(company)
            rows.extend(self._rows_from_roots(payload.get("root_domains", ""), company, source))
            rows.extend(
                self._rows_from_subdomains(
                    payload.get("subdomains", "") or payload.get("subdomain", ""),
                    payload.get("root_domains", ""),
                    company,
                    source,
                    auto_company=False,
                )
            )
            rows.extend(self._rows_from_urls(payload.get("urls", "") or payload.get("url", ""), company, source))
        elif mode == "batch":
            rows.extend(
                self._rows_from_subdomains(
                    payload.get("subdomains", ""),
                    "",
                    "",
                    source,
                    auto_company=True,
                )
            )
        elif mode == "url_extract":
            rows.extend(self._rows_from_urls(payload.get("text", "") or payload.get("urls", ""), clean_company(company), source))
        elif mode == "excel":
            rows.extend(payload.get("items") or [])
        else:
            rejected.append({"input": mode, "reason": "未知导入模式"})

        rows = dedupe_rows([row for row in rows if row])
        return {
            "items": rows,
            "rejected": rejected,
            "counts": {"valid": len(rows), "rejected": len(rejected)},
        }

    def _rows_from_roots(self, text: Any, company: str, source: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for host in extract_domains(str(text or "")):
            root = self.psl.registrable_domain(host)
            if root:
                rows.append(make_asset_row(clean_company(company), root, "", "", source))
        return rows

    def _rows_from_subdomains(
        self,
        text: Any,
        root_text: Any,
        company: str,
        source: str,
        *,
        auto_company: bool,
    ) -> list[dict[str, Any]]:
        provided_roots = [self.psl.registrable_domain(host) for host in extract_domains(str(root_text or ""))]
        provided_roots = [root for root in provided_roots if root]
        rows: list[dict[str, Any]] = []
        with self._lock, self._connect() as conn:
            for host in extract_domains(str(text or "")):
                root = match_root(host, provided_roots) or self.psl.registrable_domain(host)
                if not root:
                    continue
                subdomain = host if host != root else ""
                row_company = self._company_for_root(conn, root, company) if auto_company else clean_company(company)
                rows.append(make_asset_row(row_company, root, subdomain, "", source))
        return rows

    def _rows_from_urls(self, text: Any, company: str, source: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for url in extract_urls(str(text or "")):
            parsed = urlparse(url)
            host = normalize_host(parsed.hostname or "")
            if not is_valid_domain(host):
                continue
            root = self.psl.registrable_domain(host)
            if not root:
                continue
            subdomain = host if host != root else ""
            rows.append(make_asset_row(clean_company(company), root, subdomain, url, source))
        return rows

    def import_assets(self, items: list[dict[str, Any]], import_type: str = "manual", raw_content: str = "") -> dict[str, Any]:
        imported_ids: list[int] = []
        created = 0
        updated = 0
        failed = 0
        now = now_iso()

        with self._lock, self._connect() as conn:
            for raw in items:
                row = normalize_asset_item(raw, self.psl)
                if not row:
                    failed += 1
                    continue
                row["updated_at"] = now
                company_id = self._ensure_company(conn, row["company_name"])
                existing = conn.execute(
                    "SELECT id, source, url FROM assets WHERE domain_key = ?",
                    (row["domain_key"],),
                ).fetchone()
                if existing:
                    updated += 1
                    source = merge_source(existing["source"], row["source"])
                    conn.execute(
                        """
                        UPDATE assets
                        SET company_id = ?, company_name = ?, root_domain = ?, subdomain = ?,
                            url = COALESCE(NULLIF(?, ''), url),
                            resolved_ip = NULL, status = 'Pending', source = ?,
                            notes = ?, updated_at = ?, last_resolved_at = NULL
                        WHERE domain_key = ?
                        """,
                        (
                            company_id,
                            row["company_name"],
                            row["root_domain"],
                            row["subdomain"],
                            row["url"],
                            source,
                            row["notes"],
                            row["updated_at"],
                            row["domain_key"],
                        ),
                    )
                    imported_ids.append(int(existing["id"]))
                else:
                    created += 1
                    conn.execute(
                        """
                        INSERT INTO assets (
                            company_id, company_name, root_domain, subdomain, url, domain_key,
                            resolved_ip, status, source, notes, created_at, updated_at, last_resolved_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, NULL, 'Pending', ?, ?, ?, ?, NULL)
                        """,
                        (
                            company_id,
                            row["company_name"],
                            row["root_domain"],
                            row["subdomain"],
                            row["url"],
                            row["domain_key"],
                            row["source"],
                            row["notes"],
                            now,
                            now,
                        ),
                    )
                    imported_ids.append(int(conn.execute("SELECT last_insert_rowid()").fetchone()[0]))

            self._record_import(conn, import_type, raw_content, len(items), created + updated, failed)

        for asset_id in imported_ids:
            self.queue_resolve(asset_id)

        return {
            "imported": len(imported_ids),
            "created": created,
            "updated": updated,
            "failed": failed,
            "ids": imported_ids,
        }

    def _record_import(
        self,
        conn: sqlite3.Connection,
        import_type: str,
        raw_content: str,
        total: int,
        success: int,
        failed: int,
    ) -> None:
        conn.execute(
            """
            INSERT INTO import_records (
                import_type, raw_content, total_count, success_count, failed_count, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (import_type, raw_content[:50000], total, success, failed, now_iso()),
        )

    def list_assets(self, filters: dict[str, Any]) -> dict[str, Any]:
        page = max(1, int(filters.get("page") or 1))
        page_size = min(200, max(10, int(filters.get("page_size") or 50)))
        where: list[str] = []
        params: list[Any] = []

        for key, column in (
            ("company", "company_name"),
            ("root_domain", "root_domain"),
            ("status", "status"),
            ("source", "source"),
        ):
            value = str(filters.get(key) or "").strip()
            if value:
                where.append(f"{column} LIKE ?")
                params.append(f"%{value}%")

        keyword = str(filters.get("q") or filters.get("keyword") or "").strip()
        if keyword:
            where.append(
                "(company_name LIKE ? OR root_domain LIKE ? OR subdomain LIKE ? OR url LIKE ? "
                "OR resolved_ip LIKE ? OR source LIKE ? OR domain_key LIKE ?)"
            )
            params.extend([f"%{keyword}%"] * 7)

        clause = "WHERE " + " AND ".join(where) if where else ""
        with self._lock, self._connect() as conn:
            total = conn.execute(f"SELECT COUNT(*) FROM assets {clause}", params).fetchone()[0]
            rows = conn.execute(
                f"""
                SELECT * FROM assets
                {clause}
                ORDER BY updated_at DESC, id DESC
                LIMIT ? OFFSET ?
                """,
                [*params, page_size, (page - 1) * page_size],
            ).fetchall()
        return {
            "items": [asset_row_to_dict(row) for row in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def stats(self) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            total = conn.execute("SELECT COUNT(*) FROM assets").fetchone()[0]
            resolved = conn.execute("SELECT COUNT(*) FROM assets WHERE status = 'Resolved'").fetchone()[0]
            pending = conn.execute("SELECT COUNT(*) FROM assets WHERE status = 'Pending'").fetchone()[0]
            failed = conn.execute("SELECT COUNT(*) FROM assets WHERE status = 'Failed'").fetchone()[0]
            roots = conn.execute("SELECT COUNT(DISTINCT root_domain) FROM assets").fetchone()[0]
            companies = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        return {
            "total": total,
            "resolved": resolved,
            "pending": pending,
            "failed": failed,
            "roots": roots,
            "companies": companies,
        }

    def list_import_records(self, limit: int = 100) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT * FROM import_records
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
            total = conn.execute("SELECT COUNT(*) FROM import_records").fetchone()[0]
        return {"items": [dict(row) for row in rows], "total": total}

    def queue_resolve(self, asset_id: int) -> None:
        self._resolve_queue.put(asset_id)

    def queue_resolve_all(self) -> int:
        with self._lock, self._connect() as conn:
            rows = conn.execute("SELECT id FROM assets").fetchall()
        for row in rows:
            self.queue_resolve(int(row["id"]))
        return len(rows)

    def _resolver_loop(self) -> None:
        while True:
            asset_id = self._resolve_queue.get()
            try:
                self._resolve_one(asset_id)
            finally:
                self._resolve_queue.task_done()

    def _resolve_one(self, asset_id: int) -> None:
        with self._lock, self._connect() as conn:
            row = conn.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
        if not row:
            return
        host = row["subdomain"] or host_from_url(row["url"]) or row["root_domain"]
        ip = resolve_host(host)
        status = "Resolved" if ip else "Failed"
        with self._lock, self._connect() as conn:
            conn.execute(
                """
                UPDATE assets
                SET resolved_ip = ?, status = ?, last_resolved_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (ip or "", status, now_iso(), now_iso(), asset_id),
            )

    def extract_content(self, content: str) -> dict[str, list[str]]:
        urls = extract_urls(content)
        domains = extract_domains(content)
        for url in urls:
            host = host_from_url(url)
            if host and host not in domains:
                domains.append(host)
        ips = extract_ips(content)
        return {
            "domains": sorted(set(domains)),
            "ips": sorted(set(ips)),
            "urls": sorted(set(urls)),
        }

    def parse_excel(self, file_bytes: bytes) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment guard
            raise RuntimeError("未安装 openpyxl，无法读取 Excel 文件") from exc

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(rows_iter)]
        except StopIteration:
            return {"items": [], "counts": {"valid": 0, "rejected": 0}, "rejected": []}

        header_map = {name: idx for idx, name in enumerate(headers)}
        aliases = {
            "company": ["公司名称", "company", "company_name"],
            "root_domain": ["根域名", "root_domain", "root"],
            "subdomain": ["子域名", "subdomain", "domain"],
            "url": ["URL", "url"],
        }

        def cell(row: tuple[Any, ...], key: str) -> str:
            for alias in aliases[key]:
                idx = header_map.get(alias)
                if idx is not None and idx < len(row):
                    return str(row[idx] or "").strip()
            return ""

        raw_items: list[dict[str, Any]] = []
        rejected: list[dict[str, str]] = []
        for line_no, row in enumerate(rows_iter, start=2):
            item = {
                "company": cell(row, "company"),
                "root_domain": cell(row, "root_domain"),
                "subdomain": cell(row, "subdomain"),
                "url": cell(row, "url"),
                "source": "Excel导入",
            }
            if not any(item.values()):
                continue
            normalized = normalize_asset_item(item, self.psl)
            if normalized:
                raw_items.append(preview_row_from_normalized(normalized))
            else:
                rejected.append({"input": f"第 {line_no} 行", "reason": "无法识别有效域名"})

        items = dedupe_rows(raw_items)
        return {
            "items": items,
            "rejected": rejected,
            "counts": {"valid": len(items), "rejected": len(rejected)},
        }

    def generate_fofa_query(self, payload: dict[str, Any]) -> dict[str, Any]:
        include_domain = bool(payload.get("include_domain", True))
        include_cert = bool(payload.get("include_cert", True))
        domains = parse_domain_lines(str(payload.get("domains") or ""), self.psl)
        if not include_domain and not include_cert:
            raise ValueError("请至少选择 domain 或 cert")
        if not domains:
            raise ValueError("没有可用域名")

        clauses: list[str] = []
        for domain in domains:
            if include_domain:
                clauses.append(f'domain="{domain}"')
            if include_cert:
                clauses.append(f'cert="{domain}"')
        query_text = format_fofa_query(" || ".join(clauses))

        with self._lock, self._connect() as conn:
            conn.execute(
                """
                INSERT INTO fofa_query_history (
                    query_text, processed_count, include_domain, include_cert, created_at
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (query_text, len(domains), int(include_domain), int(include_cert), now_iso()),
            )
            record_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])

        return {
            "success": True,
            "record_id": record_id,
            "processed_count": len(domains),
            "query_text": query_text,
            "domains": domains,
        }

    def fofa_history(self, limit: int = 50) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT * FROM fofa_query_history
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            item["query_preview"] = item["query_text"][:120] + ("..." if len(item["query_text"]) > 120 else "")
            items.append(item)
        return {"items": items}

    def delete_fofa_history(self, record_id: int) -> dict[str, Any]:
        with self._lock, self._connect() as conn:
            cur = conn.execute("DELETE FROM fofa_query_history WHERE id = ?", (record_id,))
        return {"deleted": cur.rowcount}

    def export_csv(self, filters: dict[str, Any] | None = None) -> str:
        filters = dict(filters or {})
        filters["page_size"] = 200
        page = 1
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["公司名称", "根域名", "子域名", "URL", "解析IP", "状态", "来源", "更新时间"])
        while True:
            filters["page"] = page
            data = self.list_assets(filters)
            for item in data["items"]:
                writer.writerow([
                    item["company_name"],
                    item["root_domain"],
                    item.get("subdomain") or "",
                    item.get("url") or "",
                    item.get("resolved_ip") or "",
                    item["status"],
                    item["source"],
                    item["updated_at"],
                ])
            if page * data["page_size"] >= data["total"]:
                break
            page += 1
        return "\ufeff" + output.getvalue()

    def backfill_from_result(self, target: str, result_dir: Path, company: str = "") -> dict[str, Any]:
        rows: list[dict[str, Any]] = []
        phase1 = read_json(result_dir / "phase1_subdomains.json")
        if isinstance(phase1, dict):
            sources = phase1.get("sources") or {}
            used_source_map = False
            if isinstance(sources, dict):
                for source, hosts in sources.items():
                    if isinstance(hosts, list):
                        used_source_map = True
                        rows.extend(
                            self._rows_from_subdomains(
                                "\n".join(str(host) for host in hosts),
                                "",
                                company,
                                str(source),
                                auto_company=True,
                            )
                        )
            if not used_source_map and isinstance(phase1.get("subdomains"), list):
                rows.extend(
                    self._rows_from_subdomains(
                        "\n".join(str(host) for host in phase1["subdomains"]),
                        "",
                        company,
                        "scan",
                        auto_company=True,
                    )
                )

        phase2 = read_json(result_dir / "phase2_urls.json")
        if isinstance(phase2, dict):
            for source in ("gau", "katana", "all"):
                urls = phase2.get(source)
                if isinstance(urls, list):
                    rows.extend(self._rows_from_urls("\n".join(str(url) for url in urls), company, source))

        phase4 = read_json(result_dir / "phase4_fuzz.json")
        if isinstance(phase4, list):
            urls = [str(item.get("url", "")) for item in phase4 if isinstance(item, dict)]
            rows.extend(self._rows_from_urls("\n".join(urls), company, "ffuf"))

        rows = dedupe_rows(rows)
        return self.import_assets(rows, import_type="scan_backfill", raw_content=str(result_dir))


def normalize_host(value: str) -> str:
    host = str(value or "").strip().lower().rstrip(".")
    if "://" in host:
        parsed = urlparse(host)
        host = parsed.hostname or ""
    else:
        host = host.split("/")[0]
        if "@" in host:
            host = host.rsplit("@", 1)[1]
        if ":" in host:
            host = host.split(":", 1)[0]
    if host.startswith("www."):
        host = host[4:]
    return host.rstrip(".")


def normalize_url(value: str) -> str | None:
    url = str(value or "").strip().rstrip(".,;，。；、)")
    url = re.sub(r"[?？]+$", "", url)
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        return None
    host = normalize_host(parsed.hostname)
    port = f":{parsed.port}" if parsed.port else ""
    path = parsed.path or ""
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{parsed.scheme}://{host}{port}{path}{query}"


def is_valid_label(label: str) -> bool:
    if not label or len(label) > 63:
        return False
    if label.startswith("-") or label.endswith("-"):
        return False
    return bool(re.fullmatch(r"[a-z0-9-]+", label))


def is_valid_domain(host: str) -> bool:
    host = normalize_host(host)
    if len(host) > 253 or "." not in host:
        return False
    try:
        ipaddress.ip_address(host)
        return False
    except ValueError:
        pass
    return all(is_valid_label(label) for label in host.split("."))


def extract_urls(text: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for match in URL_RE.finditer(normalize_input_text(text)):
        url = normalize_url(match.group(0))
        if url and url not in seen:
            seen.add(url)
            found.append(url)
    return found


def extract_domains(text: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    clean_text = URL_RE.sub(" ", normalize_input_text(text))
    for match in DOMAIN_RE.finditer(clean_text):
        host = normalize_host(match.group(0))
        if is_valid_domain(host) and host not in seen:
            seen.add(host)
            found.append(host)
    return found


def extract_ips(text: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for match in IP_RE.finditer(normalize_input_text(text)):
        value = match.group(0)
        try:
            ipaddress.ip_address(value)
        except ValueError:
            continue
        if value not in seen:
            seen.add(value)
            found.append(value)
    return found


def normalize_input_text(text: str) -> str:
    return str(text or "").replace("\\r", " ").replace("\\n", " ")


def clean_company(value: str) -> str:
    value = str(value or "").strip()
    return value or DEFAULT_COMPANY


def source_for_mode(mode: str) -> str:
    return {
        "single": "手工录入",
        "batch": "批量导入",
        "url_extract": "URL提取",
        "excel": "Excel导入",
    }.get(mode, "手工录入")


def match_root(host: str, roots: list[str]) -> str | None:
    for root in roots:
        if host == root or host.endswith("." + root):
            return root
    return None


def make_asset_row(company: str, root_domain: str, subdomain: str, url: str, source: str) -> dict[str, Any]:
    return {
        "company": clean_company(company),
        "company_name": clean_company(company),
        "root_domain": root_domain,
        "subdomain": subdomain or "",
        "url": url or "",
        "domain_key": build_domain_key(root_domain, subdomain),
        "source": source or "手工录入",
        "status": "Pending",
        "notes": "",
    }


def normalize_asset_item(item: dict[str, Any], psl: PublicSuffix) -> dict[str, Any] | None:
    company = clean_company(str(item.get("company_name") or item.get("company") or ""))
    url = normalize_url(str(item.get("url", ""))) if item.get("url") else ""
    subdomain = normalize_host(str(item.get("subdomain", ""))) if item.get("subdomain") else ""
    root = normalize_host(str(item.get("root_domain", ""))) if item.get("root_domain") else ""

    if url and not subdomain:
        subdomain = normalize_host(urlparse(url).hostname or "")
    if subdomain and not root:
        root = psl.registrable_domain(subdomain) or ""
    if root:
        root = psl.registrable_domain(root) or root
    if not root or not is_valid_domain(root):
        return None
    if subdomain and not is_valid_domain(subdomain):
        return None
    if subdomain == root:
        subdomain = ""

    return {
        "company_name": company,
        "root_domain": root,
        "subdomain": subdomain,
        "url": url,
        "domain_key": build_domain_key(root, subdomain),
        "source": str(item.get("source") or "手工录入"),
        "notes": str(item.get("notes") or ""),
    }


def preview_row_from_normalized(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "company": item["company_name"],
        "company_name": item["company_name"],
        "root_domain": item["root_domain"],
        "subdomain": item.get("subdomain") or "",
        "url": item.get("url") or "",
        "domain_key": item["domain_key"],
        "source": item.get("source") or "手工录入",
        "status": "Pending",
        "notes": item.get("notes") or "",
    }


def build_domain_key(root: str, subdomain: str | None = "") -> str:
    return (subdomain or root or "").strip().lower()


def dedupe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for row in rows:
        key = row.get("domain_key") or build_domain_key(row["root_domain"], row.get("subdomain", ""))
        if key in seen:
            continue
        seen.add(key)
        row["domain_key"] = key
        result.append(row)
    return result


def merge_source(old: str, new: str) -> str:
    parts: list[str] = []
    for value in (old or "", new or ""):
        for item in re.split(r"[,，/]", value):
            item = item.strip()
            if item and item not in parts:
                parts.append(item)
    return ",".join(parts) or new or old


def host_from_url(url: str | None) -> str:
    if not url:
        return ""
    return normalize_host(urlparse(url).hostname or "")


def resolve_host(host: str) -> str | None:
    if not host:
        return None
    for _ in range(3):
        try:
            infos = socket.getaddrinfo(host, None, type=socket.SOCK_STREAM)
        except socket.gaierror:
            continue
        for info in infos:
            ip = info[4][0]
            if ip:
                return ip
    return None


def parse_domain_lines(raw_text: str, psl: PublicSuffix) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for host in extract_domains(raw_text):
        domain = psl.registrable_domain(host) or host
        if domain and domain not in seen:
            seen.add(domain)
            result.append(domain)
    return result


def format_fofa_query(query_text: str, max_width: int = 140) -> str:
    parts = query_text.split(" || ")
    lines: list[str] = []
    current = ""
    for part in parts:
        candidate = part if not current else f"{current} || {part}"
        if len(candidate) > max_width and current:
            lines.append(current)
            current = part
        else:
            current = candidate
    if current:
        lines.append(current)
    return "\n".join(lines)


def asset_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    data = dict(row)
    data["company"] = data.get("company_name") or data.get("company") or DEFAULT_COMPANY
    data["asset_key"] = data.get("domain_key")
    data["ip_address"] = data.get("resolved_ip")
    return data


def read_json(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
